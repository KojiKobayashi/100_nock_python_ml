# %%
'''
31th nock
'''

import pandas as pd
import glob
import os

# %%
m_store = pd.read_csv("m_store.csv")
m_area = pd.read_csv("m_area.csv")

# %%
# 対象ファイル
current_dir = os.getcwd()
tbl_order_file = os.path.join(current_dir, 'tbl_order_*.csv')
tbl_order_files = glob.glob(tbl_order_file)
tbl_order_files

# %%
order_all = pd.DataFrame()
for file in tbl_order_files:
    order_tmp = pd.read_csv(file)
    print(f'{file}:{len(order_tmp)}')
    order_all = pd.concat([order_all, order_tmp], ignore_index=True)

# %%
# 保守用店舗削除
print(len(order_all))
order_all = order_all.loc[order_all['store_id'] != 999]
print(len(order_all))

# %%
order_all = pd.merge(order_all, m_store, on='store_id', how='left')
order_all = pd.merge(order_all, m_area, on='area_cd', how='left')

# %%
# マスターにないコードに対応した文字列を設定
order_all.loc[order_all['takeout_flag']==0, 'take_out_name'] = 'デリバリー'
order_all.loc[order_all['takeout_flag']==1, 'take_out_name'] = 'お持ち帰り'

# %%
order_all.loc[order_all['status'] == 0, 'status_name'] = '受付'
order_all.loc[order_all['status'] == 1, 'status_name'] = 'お支払済'
order_all.loc[order_all['status'] == 2, 'status_name'] = 'お渡し済'
order_all.loc[order_all['status'] == 3, 'status_name'] = 'キャンセル'

# %%
order_all.loc[:, 'order_date'] = pd.to_datetime(order_all['order_accept_date']).dt.date

# %%
order_all.head()

# %%
import openpyxl

# %%
wb = openpyxl.Workbook()
ws = wb['Sheet']
ws.cell(1,1).value = '書き込みテストです。'
wb.save('test.xlsx')
wb.close()

# %%
wb = openpyxl.load_workbook('test.xlsx', read_only=True)
ws = wb['Sheet']
print(ws.cell(1,1).value)
wb.close()

# %%
store_id = 1
store_df = order_all[order_all['store_id']==store_id].copy()
store_name = store_df['store_name'].unique()[0]
store_sales_total = store_df.loc[store_df['status'].isin([1,2])]['total_amount'].sum()
store_sales_takeout = store_df.loc[store_df['status'] == 1]['total_amount'].sum()
store_sales_delivery = store_df.loc[store_df['status']==2]['total_amount'].sum()

# %%
print(f'売上額確認 {store_sales_total} = {store_sales_takeout + store_sales_delivery}')
output_df = store_df[['order_accept_date', 'customer_id', 'total_amount', 'take_out_name', 'status_name']]
output_df.head()  # ほんと出力が違う？

# %%
from openpyxl.utils.dataframe import dataframe_to_rows

store_title = f'{store_id}_{store_name}'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = store_title

ws.cell(1,1).value = f'{store_title} 売上データ'
rows = dataframe_to_rows(output_df, index=False, header=True)

row_start = 3
col_start = 2

for row_no, row in enumerate(rows, row_start):
    for col_no, value in enumerate(row, col_start):
        ws.cell(row_no, col_no).value = value

filename = f'{store_title}.xlsx'
wb.save(filename)  # 本と中身がちがう？
wb.close()

# %%
